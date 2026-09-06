"""Private analytics runtime. Launch from cil-platform; never run the anonymous launcher."""
import hashlib
import hmac
import json
import csv as stdlib_csv
import os
import sys
import time
from datetime import datetime,timezone
from pathlib import Path
from uuid import UUID
from dotenv import dotenv_values

PROJECT=Path(__file__).resolve().parents[1]
config={**dotenv_values(PROJECT/'.env'),**os.environ}
# Deliberately do not copy Supabase/admin secrets into the analytics process.
for key,value in config.items():
 if value and (key.startswith(('OPENAI_','AZURE_','ANTHROPIC_','GEMINI_','OLLAMA_','SARVAM_')) or key in ('DF_BRIDGE_SECRET','CIL_PROCESSING_ROOT','DF_SANDBOX','CIL_PRIMARY_PROVIDER','CIL_FALLBACK_PROVIDER')):os.environ[key]=value
secret=os.environ.get('DF_BRIDGE_SECRET','')
if len(secret)<32:raise SystemExit('Set DF_BRIDGE_SECRET (at least 32 random characters) in cil-platform/.env.')
processing=Path(config.get('CIL_PROCESSING_ROOT') or PROJECT/'Data'/'.processing').resolve()
os.environ.update(AUTH_PROVIDER='cil',ALLOW_ANONYMOUS='false',HOST='private',DATA_FORMULATOR_HOME=str(processing/'workspaces'),DISABLE_DATA_CONNECTORS='true',DISABLE_CUSTOM_MODELS='true',DISABLE_DISPLAY_KEYS='true',SANDBOX=config.get('DF_SANDBOX') or 'local')
sys.path.insert(0,str(PROJECT/'data-analyser'/'py-src'))
from flask import request,abort
from data_formulator.auth.providers.base import AuthProvider,AuthResult,AuthenticationError
from data_formulator.auth import providers
class CILProvider(AuthProvider):
 @property
 def name(self):return 'cil'
 def authenticate(self,request):
  if not hmac.compare_digest(request.headers.get('X-CIL-Bridge',''),secret):raise AuthenticationError('Invalid private bridge.')
  identity=request.headers.get('X-CIL-User','')
  if not identity or any(c not in 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_-' for c in identity):raise AuthenticationError('Invalid identity.')
  return AuthResult(user_id=identity)
 def get_auth_info(self):return {'action':'none'}
providers._PROVIDER_REGISTRY['cil']=CILProvider
from data_formulator.app import app
from data_formulator.auth.identity import get_identity_id
from data_formulator.workspace_factory import get_workspace,get_workspace_manager
from data_formulator.error_handler import json_ok
from data_formulator.model_registry import model_registry
from cryptography.fernet import Fernet
import base64
base_models=dict(model_registry._models)
def environment_model(role):
 default='sarvam' if role=='primary' else 'gemini'
 preferred=os.environ.get('CIL_PRIMARY_PROVIDER' if role=='primary' else 'CIL_FALLBACK_PROVIDER',default).strip().lower()
 # Custom OpenAI-compatible providers (including Sarvam) expose endpoint
 # ``openai``. Match their server-generated provider id before falling back
 # to a built-in endpoint match.
 exact=next((m for m in base_models.values() if m.get('id','').startswith(f'global-{preferred}-')),None)
 if exact:return exact
 return next((m for m in base_models.values() if m.get('endpoint')==preferred),None)
def refresh_local_models():
 path=processing/'private'/'model-providers.enc'
 entries=json.loads(Fernet(base64.urlsafe_b64encode(hashlib.sha256(secret.encode()).digest())).decrypt(path.read_bytes())) if path.exists() else []
 models=dict(base_models)
 configured_primary=next((p for p in entries if p.get('role')=='primary' and p.get('models')),None)
 configured_fallbacks=[p for p in entries if p.get('role')=='fallback' and p.get('models')]
 server_primary=environment_model('primary')
 gemini_fallback=environment_model('fallback')
 primary=server_primary or configured_primary
 if primary:
  def runtime_config(provider,name):
   return {'endpoint':provider['endpoint'],'model':name,'api_key':provider['api_key'],'api_base':provider.get('api_base',''),'api_version':provider.get('api_version',''),'timeout_seconds':provider.get('timeout_seconds',30)}
  auto=dict(primary) if server_primary else runtime_config(primary,primary['models'][0])
  fallback_configs=[]
  if gemini_fallback and gemini_fallback.get('id')!=auto.get('id'):fallback_configs.append(dict(gemini_fallback))
  fallback_configs.extend(runtime_config(p,p['models'][0]) for p in configured_fallbacks)
  auto.update({'id':'cil-auto','provider_display':'CIL Auto · server primary with Gemini fallback','fallbacks':fallback_configs,'timeout_seconds':18})
  # CIL Auto is deliberately first so a fresh workbench chooses the secured
  # server-side route instead of presenting raw provider entries as the default.
  models={'cil-auto':auto,**models}
 for provider in entries:
  if provider.get('role') in ('primary','fallback'):continue
  for name in provider['models']:
   id='cil-'+provider['id']+'-'+name
   models[id]={'id':id,'endpoint':provider['endpoint'],'model':name,'api_key':provider['api_key'],'api_base':provider['api_base'],'api_version':provider['api_version'],'provider_display':provider['name'],'timeout_seconds':provider.get('timeout_seconds',30)}
 model_registry._models=models

@app.before_request
def require_private_bridge():
 if not hmac.compare_digest(request.headers.get('X-CIL-Bridge',''),secret):return {'error':'Private analytics service'},403
 get_identity_id()
 refresh_local_models()
 if request.is_json:
  body=request.get_json(silent=True) or {}
  model=body.get('model')
  if model is not None and (not isinstance(model,dict) or not model.get('is_global') or not model_registry.get_config(model.get('id'))):
   return {'status':'error','error':{'code':'ACCESS_DENIED','message':'Select a server-configured model.'}},403

# The integrated analyst has exploration/report skills only: no data loaders or external document fetching.
import data_formulator.analyst.agent as analyst_module
original_registry=analyst_module.build_registry
def scoped_registry(*args,**kwargs):
 registry=original_registry(*args,**kwargs)
 for mapping in (registry.metas,registry.skills,registry.tool_specs):
  for key in list(mapping):
   if key not in ('core','explore','report'):mapping.pop(key,None)
 return registry
analyst_module.build_registry=scoped_registry

@app.get('/cil/health')
def health():
 models=model_registry.list_public()
 return json_ok({'models':models,'model_ready':bool(models),'sandbox':app.config['CLI_ARGS']['sandbox']})

@app.post('/cil/import')
def import_sources():
 try:return import_sources_impl()
 except (ValueError,KeyError,stdlib_csv.Error) as exc:
  return {'status':'error','error':{'message':str(exc)[:350]}},422

def import_sources_impl():
 import pyarrow as pa
 import pyarrow.csv as csv
 import pyarrow.parquet as pq
 import duckdb
 from openpyxl import load_workbook
 from data_formulator.datalake.workspace_metadata import TableMetadata,ColumnInfo
 sid=str(UUID(request.headers['X-Workspace-Id']));identity=get_identity_id();mgr=get_workspace_manager(identity)
 if not mgr.workspace_exists(sid):mgr.create_workspace(sid)
 ws=get_workspace(identity);data=request.get_json();tables=[]
 for index,source in enumerate(data['sources']):
  src=(processing/source['snapshot']).resolve()
  if not src.is_relative_to(processing/'snapshots'/sid) or not src.is_file():abort(403)
  name=f"{source['entity'].lower()}_{source['family']}_{index+1}"
  dest=ws.get_file_path(name+'.parquet');ext=src.suffix.lower()
  if ext=='.csv':
   from csv_import import csv_to_parquet
   csv_to_parquet(src,dest)
  elif ext=='.xlsx':
   book=load_workbook(src,read_only=True,data_only=True)
   if len(book.sheetnames)>1 and not source.get('sheet'):raise ValueError('Specify an Excel sheet for multi-sheet input.')
   sheet=book[source['sheet']] if source.get('sheet') else book[book.sheetnames[0]]
   rows=sheet.iter_rows(values_only=True);names=[str(x or '') for x in next(rows)]
   if not all(names) or len(set(names))!=len(names):raise ValueError('Excel requires unique non-empty headers.')
   # Write all Excel values as strings to prevent later batches being silently coerced to a sampled schema.
   schema=pa.schema([(n,pa.string()) for n in names]);batch=[]
   with pq.ParquetWriter(dest,schema,compression='zstd') as writer:
    for row in rows:
     batch.append({n:None if v is None else str(v) for n,v in zip(names,row)})
     if len(batch)>=10000:writer.write_table(pa.Table.from_pylist(batch,schema=schema));batch=[]
    if batch:writer.write_table(pa.Table.from_pylist(batch,schema=schema))
   book.close()
  elif ext=='.parquet':
   infile=pq.ParquetFile(src)
   with pq.ParquetWriter(dest,infile.schema_arrow,compression='zstd') as writer:
    for batch in infile.iter_batches(batch_size=65536):writer.write_batch(batch)
  elif ext=='.json':
   with duckdb.connect() as con:
    con.execute("SET memory_limit='512MB'")
    con.execute("COPY (SELECT * FROM read_json_auto(?)) TO ? (FORMAT PARQUET, COMPRESSION ZSTD)",[str(src),str(dest)])
  elif ext=='.md':
   text=src.read_text(encoding='utf8',errors='replace')
   blocks=[];heading='Document'
   for block in filter(None,(part.strip() for part in text.split('\n\n'))):
    if block.startswith('#'):
     heading=block.lstrip('#').strip() or heading
    else:blocks.append({'section':heading,'content':block})
   if not blocks:blocks=[{'section':heading,'content':text.strip()}]
   pq.write_table(pa.Table.from_pylist(blocks,schema=pa.schema([('section',pa.string()),('content',pa.string())])),dest,compression='zstd')
  else:raise ValueError('Unsupported structured file.')
  parquet=pq.ParquetFile(dest);schema=parquet.schema_arrow;rowcount=parquet.metadata.num_rows
  meta=TableMetadata(name=name,source_type='upload',filename=name+'.parquet',file_type='parquet',created_at=datetime.now(timezone.utc),content_hash=source['sha256'],file_size=dest.stat().st_size,row_count=rowcount,columns=[ColumnInfo(f.name,str(f.type)) for f in schema],original_name=source['name'],description=f"{source['entity']} / {source['family']} / {source['relative_path']}. SHA256 {source['sha256']}. All source rows; reporting period context: {data.get('period') or 'unspecified'}. Excel values may require explicit numeric/date conversion.")
  ws.add_table_metadata(meta)
  def type_for(t):
   if pa.types.is_integer(t):return 'integer'
   if pa.types.is_floating(t) or pa.types.is_decimal(t):return 'number'
   if pa.types.is_boolean(t):return 'boolean'
   if pa.types.is_date(t):return 'date'
   if pa.types.is_timestamp(t):return 'datetime'
   return 'string'
  tables.append({'kind':'input-table','id':name,'displayId':source['entity']+' · '+source['name'],'source':{'kind':'workspace','tableId':name},'snapshot':{'columns':[{'name':f.name,'type':type_for(f.type)} for f in schema],'rowCount':rowcount,'capturedAt':int(time.time()*1000),'contentHash':source['sha256']},'description':meta.description,'addedAt':int(time.time()*1000)})
 state={'__stateVersion':4,'activeWorkspace':{'id':sid,'displayName':data['title']},'inputTables':tables,'derivedTables':[],'charts':[],'generatedReports':[],'loadedTableNodes':[{'kind':'loaded-table','id':'loaded-'+t['id'],'tableId':t['id'],'parentNodeId':'__rootless_thread__','createdAt':int(time.time()*1000)} for t in tables]}
 mgr.save_session_state(sid,state)
 return json_ok({'tables':tables})

if __name__=='__main__':
 app.run(host=os.environ.get('DF_HOST','127.0.0.1'),port=int(os.environ.get('DF_PORT','5567')),threaded=True,debug=False,use_reloader=False)
